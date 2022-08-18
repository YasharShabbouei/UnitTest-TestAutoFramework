import openpyxl

def read_data_from_excel(file_address, sheetName):
    wb = openpyxl.load_workbook(filename=file_address)
    sh = wb[sheetName]
    row_max = sh.max_row
    col_max = sh.max_column
    datalist =[]

    for i in range (1, row_max+1):
        row = []
        for j in range (1, col_max+1):
            row.append(sh.cell(row=i, column=j).value)
        datalist.append(row)
    return datalist